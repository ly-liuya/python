import openpyxl
import datetime
# 加载一个工作簿
wb = openpyxl.load_workbook("./data/演示2.xlsx")

print(wb.sheetnames)

# 获取工作表
sheet = wb.worksheets[0]
#获取单元格范围
print(sheet.dimensions)

# 获取工作表的行数和列数
print(sheet.max_row,sheet.max_column)

# 获取指定单元格的值
print(sheet['c3'].value)

# 获取多个单元格（嵌套元组）
print(sheet['A2:C5'])

# 读取单元格中所有的数据
for row_ch in range(2,sheet.max_row+1):
    for col_ch in 'ABC':
        value = sheet[f"{col_ch}{row_ch}"].value
        print(value,end="\t")
    print()


