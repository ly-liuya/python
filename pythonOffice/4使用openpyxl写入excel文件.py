import openpyxl
import random

# 第一步：创建工作簿
wb = openpyxl.Workbook()

# 第二步：添加工作表
sheet= wb.active
sheet.title = "笔试成绩"

titles = ["姓名","骑马","射箭","举重"]
for col_index,title in enumerate(titles):
    sheet.cell(1,col_index+1,title)
names = ["魏延","吕布","许褚","夏侯惇","黄盖"]

for row_index,name in enumerate(names):
    sheet.cell(row_index+2,1,name)
    for col_index in range(2,5):
        sheet.cell(row_index+2,col_index,random.randrange(50,101))


# 保存工作簿
wb.save("./data/三国擂台赛成绩.xlsx")


